import openpyxl
import pandas as pd
file = "E6.xlsx"  #要处理的文件路径
wb = openpyxl.load_workbook(file)  #加载文件
data=pd.DataFrame(pd.read_excel(file))
sheet = wb.sheetnames
ws=wb[sheet[0]]
ws1=wb[sheet[1]]
ws2=wb[sheet[2]]


#进行的是进行空格的删除
for i in range(1,ws.max_row+1):
    for j in range(1,ws.max_column+1):
        old = str(ws.cell(i, j).value)
        if old is not None:
            ws.cell(i, j).value=old.strip().replace(" ",'')


#对s1表中的第3列的数据进行修改，去掉 “记-”
for i in range(1,ws.max_row+1):
   ws.cell(i,3,ws.cell(i,3).value.replace("记-",""))

#对s3表的数据和s2表的数据进行对比，如果 “工单号”和“ERPCO号”都相等，认为记录相同，添加备注列，如果相等内容为1，如果没有找到相等的内容为2.
for i in range(2, ws1.max_row + 1):
    flag = 0
    for j in range(2, ws2.max_row + 1):
        if f'{ws2.cell(i, 2).value}{ws2.cell(i, 3).value}' == f'{ws1.cell(j, 3).value}{ws1.cell(j, 4).value}':
            flag = 1
            break
        elif ws2.cell(i, 2).value != ws1.cell(j, 3).value and ws2.cell(i, 3).value != ws1.cell(j, 4).value:
            flag = 2
    if flag == 1:
        ws2.cell(i, 6, 1)
    elif flag == 2:
        ws2.cell(i, 6, 2)
wb.save(file)
wb.close()
print("处理完成")
